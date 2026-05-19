
# -*- coding: utf-8 -*-
"""
专项检查：reading 值统计、chinese 为数字问题的本质分析
"""
import pandas as pd
import json
import sys
from openpyxl import load_workbook
from collections import Counter

sys.stdout.reconfigure(encoding='utf-8')

BASE = r"d:\AgentConstruction\workspace\japanese-vocabulary-main"

wb = load_workbook(f"{BASE}\\日语单词.xlsx")
sn = wb.sheetnames[0]
ws = wb[sn]
rows = list(ws.iter_rows(values_only=True))
header = rows[0]
df = pd.DataFrame(rows[1:], columns=header)
df.rename(columns={
    header[0]: 'kana',
    header[1]: 'reading',
    header[2]: 'kanji',
    header[3]: 'chinese',
    header[4]: 'note',
}, inplace=True)

# JSON
with open(f"{BASE}\\word_data.json", "r", encoding="utf-8") as f:
    jdata = json.load(f)
jdf = pd.DataFrame(jdata)

# ─── reading 字段含义解析 ─────────────────────────────
print("=" * 70)
print("【J】reading 字段所有值的统计与样本")
print("=" * 70)
rc = Counter(df['reading'].tolist())
for val in sorted(rc.keys(), key=lambda x: (x is None, float(str(x).split('/')[0]) if x is not None and str(x).replace('.','').replace('/','').isdigit() else 999)):
    cnt = rc[val]
    if val is None:
        sample = df[df['reading'].isna()]['kana'].head(5).tolist()
    else:
        sample = df[df['reading'] == val]['kana'].head(5).tolist()
    print(f"  reading={val}: {cnt} 条 | 示例: {sample}")

# ─── chinese 为数字的行 —完整展示 ────────────────────────
print("\n" + "=" * 70)
print("【E2】chinese 字段为数字的行（即 '读音音拍数' 字段串位了）")
print("=" * 70)

def is_numeric(val):
    if pd.isna(val):
        return False
    try:
        float(str(val).strip())
        return True
    except:
        return False

numeric_cn = df[df['chinese'].apply(is_numeric)].copy()
print(f"  总计: {len(numeric_cn)} 行")
print(f"\n  chinese 数字值分布:")
print(Counter(numeric_cn['chinese'].astype(str).tolist()))

print(f"\n  前30行示例:")
print(numeric_cn[['kana','reading','kanji','chinese']].head(30).to_string())

# 分析：这些行的 reading 字段值是什么？
print(f"\n  这些行的 reading 字段值分布:")
print(Counter(numeric_cn['reading'].tolist()))

# 查 JSON 中对应的条目，看 chinese 字段是否正确
print(f"\n  对比 JSON：前20条")
jmap = {}
for _, jrow in jdf.iterrows():
    key = (str(jrow['japanese']), str(jrow['kanji']))
    jmap[key] = jrow['chinese']

correct_count = 0
wrong_count = 0
missing_count = 0
for _, row in numeric_cn.head(30).iterrows():
    key = (str(row['kana']), str(row['kanji']))
    jcn = jmap.get(key, '<<不在JSON中>>')
    status = "✓" if str(jcn) == str(row['chinese']) else "✗"
    print(f"  {status} kana={row['kana']}, kanji={row['kanji']}")
    print(f"       Excel chinese={row['chinese']}, JSON chinese={jcn}")

# ─── 单行原始数据查看（Excel A/B/C/D 列） ─────────────────
print("\n" + "=" * 70)
print("【K】查看原始数据：chinese 为数字的行，其原始全行内容")
print("=" * 70)
# 展示前10行的全部有效列
print(numeric_cn[['kana','reading','kanji','chinese','note']].head(20).to_string())

# ─── #NAME? 错误行 ────────────────────────────────────
print("\n" + "=" * 70)
print("【H2】#NAME? 错误行详细")
print("=" * 70)
for col in ['kana','kanji','chinese']:
    err_mask = df[col].astype(str).str.contains(r'#NAME\?', na=False)
    if err_mask.sum():
        print(f"  列 '{col}' 中含 #NAME? 的行:")
        print(df[err_mask][['kana','reading','kanji','chinese']].to_string())

# 搜索所有列的全部 Excel 错误字符
print("\n  全列扫描 #NAME?:")
for col in df.columns:
    if df[col].dtype == object or str(df[col].dtype) == 'object':
        mask = df[col].astype(str).str.contains(r'#NAME\?', na=False)
        if mask.sum() > 0:
            print(f"    列'{col}': {mask.sum()}行")
            print(df[mask][['kana','reading','kanji','chinese']].to_string())

# ─── 汉字列含有 ? 作为占位符的行 ──────────────────────────
print("\n" + "=" * 70)
print("【B2】kanji 列 '?' 占位符统计（?字 格式，表示汉字前缀/后缀）")
print("=" * 70)
# 以 ? 开头的 kanji（前缀型：如 ?月、?円）
prefix_q = df['kanji'].str.match(r'^\?', na=False)
print(f"  kanji 以 '?' 开头（前缀型）: {int(prefix_q.sum())} 行")
# 以 ? 结尾
suffix_q = df['kanji'].str.match(r'.*\?$', na=False)
print(f"  kanji 以 '?' 结尾（后缀型）: {int(suffix_q.sum())} 行")
# 中间含 ?
mid_q = df['kanji'].str.contains(r'\?', na=False) & ~prefix_q & ~suffix_q
print(f"  kanji 中间含 '?'（中缀型）: {int(mid_q.sum())} 行")
# 总计
total_q = df['kanji'].str.contains(r'\?', na=False)
print(f"  kanji 含 '?' 总计: {int(total_q.sum())} 行")

# ─── 整体统计汇总 ─────────────────────────────────────
print("\n" + "=" * 70)
print("【总结】问题类型统计")
print("=" * 70)
issues = {
    "reading 为空（无音拍数）": int(df['reading'].isna().sum()),
    "reading 非 0/1（2-8 表示音拍数）": int(df['reading'].apply(lambda x: x not in (0.0, 1.0) and pd.notna(x)).sum()),
    "reading 特殊值 '0/2'": int((df['reading'] == '0/2').sum()),
    "chinese 字段误存了音拍数（数字）": len(numeric_cn),
    "chinese 字段为空": int(df['chinese'].isna().sum()),
    "kanji/kana/chinese 含 '?'（不确定标记）": int((df['kana'].str.contains(r'\?', na=False) | df['kanji'].str.contains(r'\?', na=False) | df['chinese'].str.contains(r'\?', na=False)).sum()),
    "kana+chinese 重复行": int(df.duplicated(subset=['kana','chinese'], keep=False).sum()),
    "完整重复行": int(df.duplicated(subset=['kana','reading','kanji','chinese'], keep=False).sum()),
    "包含 #NAME? 错误": int((df['chinese'].astype(str).str.contains(r'#NAME\?', na=False)).sum()),
    "Excel 比 JSON 多的行数": len(set(zip(df['kana'].fillna(''), df['kanji'].fillna(''), df['chinese'].fillna(''))) - set(zip(jdf['japanese'].fillna(''), jdf['kanji'].fillna(''), jdf['chinese'].fillna('')))),
}
for k, v in issues.items():
    print(f"  {v:5d}  {k}")
