
# -*- coding: utf-8 -*-
import openpyxl
import pandas as pd
import json
import re
import sys
from openpyxl import load_workbook

# 强制 stdout utf-8
sys.stdout.reconfigure(encoding='utf-8')

BASE = r"d:\AgentConstruction\workspace\japanese-vocabulary-main"
OUT  = r"d:\AgentConstruction\workspace\japanese-vocabulary-main\analysis_report.txt"

lines = []

def p(s=""):
    lines.append(str(s))
    print(s)

# ─────────────────────────────────────────────
# 1. 读取两个 Excel
# ─────────────────────────────────────────────
p("=" * 70)
p("【1】读取 日语单词.xlsx")
p("=" * 70)

wb_main = load_workbook(f"{BASE}\\日语单词.xlsx")
p(f"Sheets: {wb_main.sheetnames}")

dfs_main = {}
for sn in wb_main.sheetnames:
    ws = wb_main[sn]
    rows = list(ws.iter_rows(values_only=True))
    p(f"\n--- Sheet: {sn} | 最大行={ws.max_row}, 最大列={ws.max_column}")
    p(f"  表头: {rows[0] if rows else 'empty'}")
    for i, r in enumerate(rows[1:8], 1):
        p(f"  第{i}行: {list(r)}")
    if rows:
        df = pd.DataFrame(rows[1:], columns=rows[0])
        dfs_main[sn] = df
        p(f"  总数据行: {len(df)}")

p()
p("=" * 70)
p("【2】读取 日语单词 copy.xlsx")
p("=" * 70)

wb_copy = load_workbook(f"{BASE}\\日语单词 copy.xlsx")
p(f"Sheets: {wb_copy.sheetnames}")

dfs_copy = {}
for sn in wb_copy.sheetnames:
    ws = wb_copy[sn]
    rows_c = list(ws.iter_rows(values_only=True))
    p(f"\n--- Sheet: {sn} | 最大行={ws.max_row}, 最大列={ws.max_column}")
    p(f"  表头: {rows_c[0] if rows_c else 'empty'}")
    for i, r in enumerate(rows_c[1:8], 1):
        p(f"  第{i}行: {list(r)}")
    if rows_c:
        df_c = pd.DataFrame(rows_c[1:], columns=rows_c[0])
        dfs_copy[sn] = df_c
        p(f"  总数据行: {len(df_c)}")

p()
p("=" * 70)
p("【3】读取 word_data.json")
p("=" * 70)
with open(f"{BASE}\\word_data.json", "r", encoding="utf-8") as f:
    word_data = json.load(f)

if isinstance(word_data, list):
    p(f"JSON 是列表，共 {len(word_data)} 条")
    if word_data and isinstance(word_data[0], dict):
        p(f"字段 keys: {list(word_data[0].keys())}")
        for item in word_data[:5]:
            p(f"  {item}")
elif isinstance(word_data, dict):
    p(f"JSON 是字典，keys: {list(word_data.keys())}")

p()
p("=" * 70)
p("【4】对比两个 Excel 文件差异")
p("=" * 70)
for sn in wb_main.sheetnames:
    if sn in dfs_copy:
        df_m = dfs_main[sn]
        df_c = dfs_copy[sn]
        p(f"\nSheet: {sn}")
        p(f"  main 行数={len(df_m)}, copy 行数={len(df_c)}")
        cols_m = set(str(c) for c in df_m.columns)
        cols_c = set(str(c) for c in df_c.columns)
        if cols_m != cols_c:
            p(f"  列差异 - main独有: {cols_m - cols_c}, copy独有: {cols_c - cols_m}")
        else:
            p(f"  列完全相同: {list(df_m.columns)}")
        min_rows = min(len(df_m), len(df_c))
        diff_rows = []
        for i in range(min_rows):
            if list(df_m.iloc[i]) != list(df_c.iloc[i]):
                diff_rows.append(i)
        if diff_rows:
            p(f"  有差异的行数: {len(diff_rows)}, 行索引(前20): {diff_rows[:20]}")
            for idx in diff_rows[:10]:
                p(f"    行{idx+2}(Excel) main: {list(df_m.iloc[idx])}")
                p(f"    行{idx+2}(Excel) copy: {list(df_c.iloc[idx])}")
        else:
            if len(df_m) != len(df_c):
                p(f"  前{min_rows}行相同，但行数不同")
            else:
                p(f"  ✓ 两文件内容完全相同（都有 {len(df_m)} 行数据）")
    else:
        p(f"Sheet {sn} 在 copy 中不存在")

# ─────────────────────────────────────────────
# 5. 详细数据检查
# ─────────────────────────────────────────────
p()
p("=" * 70)
p("【5】数据质量检查")
p("=" * 70)

# 以 main 为主进行检查
sn = wb_main.sheetnames[0]
df = dfs_main[sn].copy()
p(f"\nSheet: {sn}, 数据行数: {len(df)}")
p(f"列名: {list(df.columns)}")

# 打印更多列名（有些可能是 None/空）
p(f"列名(含空): {list(df.columns)}")
p(f"\n前10行数据:")
for i in range(min(10, len(df))):
    p(f"  row{i+1}: {list(df.iloc[i])}")

# 统计空值
p("\n--- 空值统计 ---")
for col in df.columns:
    null_count = df[col].isna().sum()
    if null_count > 0:
        p(f"  列 '{col}': {null_count} 个空值 ({100*null_count/len(df):.1f}%)")

# 重复数据
p("\n--- 重复行统计 ---")
dupes = df.duplicated()
p(f"  完整重复行数: {dupes.sum()}")
# 取第一列作为 word key
first_col = df.columns[0]
if first_col is not None:
    dupes_col = df[first_col].duplicated(keep=False)
    dupe_words = df[dupes_col][first_col].dropna().unique()
    p(f"  '{first_col}' 列重复值: {len(dupe_words)} 个")
    if len(dupe_words) > 0:
        p(f"  重复词汇示例(前20): {list(dupe_words[:20])}")

# 打印全列数据看看有几列有意义
p("\n--- 各列非空值数量 ---")
for col in df.columns:
    non_null = df[col].notna().sum()
    p(f"  列 '{col}': {non_null} 个非空值")

# 保存报告
with open(OUT, "w", encoding="utf-8") as f:
    f.write("\n".join(lines))
p(f"\n报告已保存至: {OUT}")
