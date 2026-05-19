
# -*- coding: utf-8 -*-
"""
查看关键行详情：两个 NULL chinese 行、#NAME? 行、_1 非空行
"""
import pandas as pd
import sys
from openpyxl import load_workbook
from collections import Counter

sys.stdout.reconfigure(encoding='utf-8')

BASE = r"d:\AgentConstruction\workspace\japanese-vocabulary-main"

wb = load_workbook(f"{BASE}\\日语单词.xlsx")
sn = wb.sheetnames[0]
ws = wb[sn]
rows = list(ws.iter_rows(values_only=True))
header = rows[0]
df = pd.DataFrame(rows[1:], columns=header)
df.rename(columns={
    header[0]: 'kana', header[1]: 'reading', header[2]: 'kanji',
    header[3]: 'chinese', header[4]: 'note',
}, inplace=True)

print("=" * 60)
print("【1】两个 NULL chinese 行的上下文")
print("=" * 60)
idx1 = df[df['chinese'].isna()].index.tolist()
for idx in idx1:
    start = max(0, idx-3)
    end = min(len(df), idx+4)
    print(f"\n  NULL 行索引={idx}（Excel 行 {idx+2}），上下文:")
    for i in range(start, end):
        marker = " ← NULL" if i == idx else ""
        print(f"    row{i+2}: {list(df.iloc[i][:5])}{marker}")

print("\n" + "=" * 60)
print("【2】#NAME? 行的上下文")
print("=" * 60)
err_idx = df[df['chinese'].astype(str).str.contains(r'#NAME\?', na=False)].index.tolist()
for idx in err_idx:
    start = max(0, idx-3)
    end = min(len(df), idx+4)
    print(f"\n  #NAME? 行索引={idx}（Excel 行 {idx+2}），上下文:")
    for i in range(start, end):
        marker = " ← #NAME?" if i == idx else ""
        print(f"    row{i+2}: {list(df.iloc[i][:5])}{marker}")

print("\n" + "=" * 60)
print("【3】_1 非空列的全部内容")
print("=" * 60)
note_notnull = df[df['note'].notna()]
print(f"  note 非空行数: {len(note_notnull)}")
print(note_notnull[['kana','reading','kanji','chinese','note']].to_string())

print("\n" + "=" * 60)
print("【4】reading='0/2' 的行")
print("=" * 60)
special = df[df['reading'] == '0/2']
print(special[['kana','reading','kanji','chinese']].to_string())

print("\n" + "=" * 60)
print("【5】kana+chinese 重复行详情（可能是真实重复错误）")
print("=" * 60)
dup = df.duplicated(subset=['kana','chinese'], keep=False)
dup_df = df[dup].sort_values(['kana','chinese'])
print(dup_df[['kana','reading','kanji','chinese']].to_string())

print("\n" + "=" * 60)
print("【6】所有列的 nan 统计（最后10列验证是否真全空）")
print("=" * 60)
for col in df.columns:
    non_null = df[col].notna().sum()
    print(f"  列'{col}': {non_null} 非空")

print("\n" + "=" * 60)
print("【7】まえ 附近所有条目（查看前后文）")
print("=" * 60)
mae_idx = df[df['kana'] == 'まえ'].index.tolist()
for idx in mae_idx:
    print(f"  idx={idx}(Excel {idx+2}): {list(df.iloc[idx][:5])}")
