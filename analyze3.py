
# -*- coding: utf-8 -*-
"""
全面检查日语单词 Excel，找出所有数据问题。
"""
import openpyxl
import pandas as pd
import json
import re
import sys
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding='utf-8')

BASE = r"d:\AgentConstruction\workspace\japanese-vocabulary-main"

# ── 读取数据 ──────────────────────────────────────────────
wb = load_workbook(f"{BASE}\\日语单词.xlsx")
sn = wb.sheetnames[0]
ws = wb[sn]
rows = list(ws.iter_rows(values_only=True))
header = rows[0]
df = pd.DataFrame(rows[1:], columns=header)

# 重命名前4个有效列为统一名称
df.rename(columns={
    header[0]: 'kana',        # 語 (平假名读音)
    header[1]: 'reading',     # 發音 (0/1 标记)
    header[2]: 'kanji',       # 漢字、原文
    header[3]: 'chinese',     # 中文意思
    header[4]: 'note',        # _1 备注
}, inplace=True)

# 读取 JSON
with open(f"{BASE}\\word_data.json", "r", encoding="utf-8") as f:
    jdata = json.load(f)
jdf = pd.DataFrame(jdata)

print("=" * 70)
print(f"Excel 数据行数: {len(df)}, JSON 数据行数: {len(jdf)}")
print("Excel 前10行:")
print(df[['kana','reading','kanji','chinese']].head(10).to_string())

# ══════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print("【A】空值检查")
print("=" * 70)

key_cols = ['kana', 'reading', 'kanji', 'chinese']
for col in key_cols:
    null_mask = df[col].isna()
    null_count = int(null_mask.sum())
    if null_count:
        print(f"\n  列 '{col}': {null_count} 个空值")
        sample = df[null_mask][['kana','reading','kanji','chinese']].head(10)
        print(sample.to_string(index=True))

# reading 为空的行
reading_null = df[df['reading'].isna()]
print(f"\n  'reading' 空值行数: {len(reading_null)}")
print(f"  前20行 kana: {list(reading_null['kana'].head(20))}")

# ══════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print("【B】'?' 字符（不确定条目）")
print("=" * 70)
question_mask = df['kana'].str.contains(r'\?', na=False) | \
                df['kanji'].str.contains(r'\?', na=False) | \
                df['chinese'].str.contains(r'\?', na=False)
qdf = df[question_mask]
print(f"  包含 '?' 的行数: {len(qdf)}")
print(qdf[['kana','reading','kanji','chinese']].to_string())

# ══════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print("【C】重复数据检查")
print("=" * 70)

# 完整重复行
full_dupes = df.duplicated(subset=['kana','reading','kanji','chinese'], keep=False)
full_dupe_count = int(full_dupes.sum())
print(f"\n  完整重复行（4列都相同）: {full_dupe_count} 行")
if full_dupe_count:
    duped = df[full_dupes].sort_values('kana')
    print(duped[['kana','reading','kanji','chinese']].head(20).to_string())

# kana+chinese 重复（同一个词有相同释义的两行）
kana_cn_dup = df.duplicated(subset=['kana','chinese'], keep=False)
cnt = int(kana_cn_dup.sum())
print(f"\n  kana+chinese 重复行: {cnt} 行")
if cnt:
    print(df[kana_cn_dup].sort_values('kana')[['kana','reading','kanji','chinese']].head(20).to_string())

# ══════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print("【D】reading 字段取值分析")
print("=" * 70)
print(f"  reading 唯一值: {sorted(df['reading'].dropna().unique())}")
# reading 不是 0/1/NaN 的行
invalid_reading = df[~df['reading'].isin([0.0, 1.0]) & df['reading'].notna()]
print(f"  reading 非 0/1 的行数: {len(invalid_reading)}")
if len(invalid_reading):
    print(invalid_reading[['kana','reading','kanji','chinese']].head(20).to_string())

# ══════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print("【E】kana 列格式检查（是否全为平假名/片假名/括号/汉字）")
print("=" * 70)

# 检查 kana 中出现拉丁字母的行
latin_mask = df['kana'].str.contains(r'[a-zA-Z]', na=False, regex=True)
print(f"  kana 列含拉丁字母的行数: {int(latin_mask.sum())}")
if int(latin_mask.sum()) > 0:
    print(df[latin_mask][['kana','reading','kanji','chinese']].head(20).to_string())

# kana 为空字符串
empty_str_mask = df['kana'].str.strip().eq('') if df['kana'].dtype == object else pd.Series([False]*len(df))
print(f"  kana 列为空字符串的行数: {int(empty_str_mask.sum())}")

# kana 含数字
digit_mask = df['kana'].str.contains(r'\d', na=False, regex=True)
print(f"  kana 列含数字的行数: {int(digit_mask.sum())}")
if int(digit_mask.sum()):
    print(df[digit_mask][['kana','reading','kanji','chinese']].head(20).to_string())

# ══════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print("【F】kanji 列格式检查")
print("=" * 70)

# kanji 与 kana 完全相同的行（无汉字形式，正常情况）
same_mask = df['kana'] == df['kanji']
print(f"  kanji == kana 的行数（无汉字形式，正常）: {int(same_mask.sum())}")

# kanji 为空
kanji_null = df['kanji'].isna()
print(f"  kanji 为空的行数: {int(kanji_null.sum())}")

# kanji 含拉丁字母
kanji_latin = df['kanji'].str.contains(r'[a-zA-Z]', na=False, regex=True)
print(f"  kanji 含拉丁字母的行数: {int(kanji_latin.sum())}")
if int(kanji_latin.sum()):
    print(df[kanji_latin][['kana','reading','kanji','chinese']].head(20).to_string())

# ══════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print("【G】chinese 列内容检查")
print("=" * 70)

# 含日语假名（可能误放了日语）
hiragana_in_cn = df['chinese'].str.contains(r'[぀-ゟ゠-ヿ]', na=False, regex=True)
print(f"  chinese 含日语假名的行数: {int(hiragana_in_cn.sum())}")
if int(hiragana_in_cn.sum()):
    print(df[hiragana_in_cn][['kana','reading','kanji','chinese']].to_string())

# chinese 为纯英文
english_only = df['chinese'].str.match(r'^[a-zA-Z\s\-.,!?]+$', na=False)
print(f"  chinese 为纯英文的行数: {int(english_only.sum())}")
if int(english_only.sum()):
    print(df[english_only][['kana','reading','kanji','chinese']].head(20).to_string())

# chinese 含 '?'
cn_q = df['chinese'].str.contains(r'\?', na=False)
print(f"  chinese 含 '?' 的行数: {int(cn_q.sum())}")

# ══════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print("【H】note(_1) 列的非空值")
print("=" * 70)
note_notnull = df['note'].notna()
print(f"  note 非空行数: {int(note_notnull.sum())}")
if int(note_notnull.sum()):
    print(df[note_notnull][['kana','reading','kanji','chinese','note']].to_string())

# ══════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print("【I】Excel 与 JSON 的差异（多/少的行）")
print("=" * 70)

# 用 kana+kanji+chinese 三列做 key 比较
excel_keys = set(zip(df['kana'].fillna(''), df['kanji'].fillna(''), df['chinese'].fillna('')))
json_keys  = set(zip(jdf['japanese'].fillna(''), jdf['kanji'].fillna(''), jdf['chinese'].fillna('')))

only_in_excel = excel_keys - json_keys
only_in_json  = json_keys  - excel_keys

print(f"  仅在 Excel 中存在（{len(only_in_excel)} 条）:")
for k in sorted(only_in_excel)[:30]:
    print(f"    kana={k[0]}, kanji={k[1]}, cn={k[2]}")

print(f"\n  仅在 JSON 中存在（{len(only_in_json)} 条）:")
for k in sorted(only_in_json)[:30]:
    print(f"    kana={k[0]}, kanji={k[1]}, cn={k[2]}")

# ══════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print("【J】格式不一致：括号形式检查")
print("=" * 70)

# 混用全角/半角括号
full_paren = df['kana'].str.contains(r'[（）]', na=False)
half_paren = df['kana'].str.contains(r'[()]', na=False)
both_paren = full_paren & half_paren
print(f"  kana 同时含全角和半角括号的行数: {int(both_paren.sum())}")
if int(both_paren.sum()):
    print(df[both_paren][['kana','reading','kanji','chinese']].head(20).to_string())

# 只含半角括号
only_half = ~full_paren & half_paren
print(f"  kana 只含半角括号（不含全角）的行数: {int(only_half.sum())}")
if int(only_half.sum()):
    print(df[only_half][['kana','reading','kanji','chinese']].to_string())

# kanji 中的括号情况
kanji_full = df['kanji'].str.contains(r'[（）]', na=False)
kanji_half = df['kanji'].str.contains(r'[()]', na=False)
kanji_only_half = ~kanji_full & kanji_half
print(f"  kanji 只含半角括号的行数: {int(kanji_only_half.sum())}")
if int(kanji_only_half.sum()):
    print(df[kanji_only_half][['kana','reading','kanji','chinese']].to_string())

# ══════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print("【K】中文繁简体混用检查（粗略）")
print("=" * 70)

# 典型繁体字符：這/這、爲/為、來/來、時/時
trad_chars = r'[這為來時們體實際學習國際語文動詞裏語義對應於問題發現關係]'
trad_mask = df['chinese'].str.contains(trad_chars, na=False)
print(f"  chinese 含常见繁体字的行数: {int(trad_mask.sum())}")
# 这只是粗略检测，先展示一些
print(f"  前10个繁体示例:")
print(df[trad_mask][['kana','chinese']].head(10).to_string())

# ══════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print("【L】kana 排序验证（应按五十音顺序）")
print("=" * 70)
# 检查是否大致按 あいうえお 顺序排列
kana_list = df['kana'].dropna().tolist()
out_of_order = []
for i in range(1, len(kana_list)):
    if kana_list[i] < kana_list[i-1]:
        # 可能是合法的（同音多义词排在一起），只记录跨行明显乱序
        if kana_list[i][:1] < kana_list[i-1][:1]:
            out_of_order.append((i+2, kana_list[i-1], kana_list[i]))  # Excel行号

print(f"  明显排序错误（首字母倒退）的位置数: {len(out_of_order)}")
for row_n, prev, curr in out_of_order[:20]:
    print(f"    Excel行{row_n}: prev='{prev}' → curr='{curr}'")

print("\n" + "=" * 70)
print("全部检查完成！")
print("=" * 70)
