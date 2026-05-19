
# -*- coding: utf-8 -*-
import openpyxl
import pandas as pd
import json
import re
from openpyxl import load_workbook
from copy import copy

BASE = r"d:\AgentConstruction\workspace\japanese-vocabulary-main"

# ─────────────────────────────────────────────
# 1. 读取两个 Excel 文件结构
# ─────────────────────────────────────────────
print("=" * 70)
print("【1】读取  日语单词.xlsx")
print("=" * 70)

wb_main = load_workbook(f"{BASE}\\日语单词.xlsx")
print(f"Sheets: {wb_main.sheetnames}")

dfs_main = {}
for sheet_name in wb_main.sheetnames:
    ws = wb_main[sheet_name]
    print(f"\n--- Sheet: {sheet_name} | 最大行={ws.max_row}, 最大列={ws.max_column}")
    rows = list(ws.iter_rows(values_only=True))
    print(f"  表头: {list(rows[0]) if rows else 'empty'}")
    for i, r in enumerate(rows[1:6], 1):
        print(f"  第{i}行: {list(r)}")
    # 转成 DataFrame
    if rows:
        df = pd.DataFrame(rows[1:], columns=rows[0])
        dfs_main[sheet_name] = df
        print(f"  总数据行数: {len(df)}")

print("\n" + "=" * 70)
print("【2】读取  日语单词 copy.xlsx")
print("=" * 70)

wb_copy = load_workbook(f"{BASE}\\日语单词 copy.xlsx")
print(f"Sheets: {wb_copy.sheetnames}")

dfs_copy = {}
for sheet_name in wb_copy.sheetnames:
    ws = wb_copy[sheet_name]
    print(f"\n--- Sheet: {sheet_name} | 最大行={ws.max_row}, 最大列={ws.max_column}")
    rows_c = list(ws.iter_rows(values_only=True))
    print(f"  表头: {list(rows_c[0]) if rows_c else 'empty'}")
    for i, r in enumerate(rows_c[1:6], 1):
        print(f"  第{i}行: {list(r)}")
    if rows_c:
        df_c = pd.DataFrame(rows_c[1:], columns=rows_c[0])
        dfs_copy[sheet_name] = df_c
        print(f"  总数据行数: {len(df_c)}")

print("\n" + "=" * 70)
print("【3】读取  word_data.json")
print("=" * 70)
with open(f"{BASE}\\word_data.json", "r", encoding="utf-8") as f:
    word_data = json.load(f)

if isinstance(word_data, list):
    print(f"JSON 是列表，共 {len(word_data)} 条")
    if word_data:
        print(f"第1条 keys: {list(word_data[0].keys()) if isinstance(word_data[0], dict) else 'not dict'}")
        for item in word_data[:3]:
            print(f"  {item}")
elif isinstance(word_data, dict):
    print(f"JSON 是字典，keys: {list(word_data.keys())}")
    for k in list(word_data.keys())[:3]:
        print(f"  {k}: {word_data[k]}")

print("\n" + "=" * 70)
print("【4】对比两个 Excel 文件差异")
print("=" * 70)

for sheet_name in wb_main.sheetnames:
    if sheet_name in dfs_copy:
        df_m = dfs_main[sheet_name]
        df_c = dfs_copy[sheet_name]
        print(f"\nSheet: {sheet_name}")
        print(f"  main 行数={len(df_m)}, copy 行数={len(df_c)}")
        # 列差异
        cols_m = set(df_m.columns)
        cols_c = set(df_c.columns)
        if cols_m != cols_c:
            print(f"  列差异 - main独有: {cols_m - cols_c}, copy独有: {cols_c - cols_m}")
        else:
            print(f"  列完全相同: {list(df_m.columns)}")
        # 找不同行
        min_rows = min(len(df_m), len(df_c))
        diff_rows = []
        for i in range(min_rows):
            row_m = df_m.iloc[i]
            row_c = df_c.iloc[i]
            if not row_m.equals(row_c):
                diff_rows.append(i)
        if diff_rows:
            print(f"  有差异的行数: {len(diff_rows)}, 行索引(前10): {diff_rows[:10]}")
            for idx in diff_rows[:5]:
                print(f"    行{idx+1} main: {list(df_m.iloc[idx])}")
                print(f"    行{idx+1} copy: {list(df_c.iloc[idx])}")
        else:
            if len(df_m) != len(df_c):
                print(f"  前{min_rows}行相同，但行数不同")
            else:
                print(f"  两文件内容完全相同")
    else:
        print(f"Sheet {sheet_name} 在 copy 中不存在")
