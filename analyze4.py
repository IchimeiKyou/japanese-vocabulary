
# -*- coding: utf-8 -*-
"""
专注检查 reading 字段、chinese 为数字、错误内容等特殊问题
"""
import pandas as pd
import json
import sys
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding='utf-8')

BASE = r"d:\AgentConstruction\workspace\japanese-vocabulary-main"

wb = load_workbook(f"{BASE}\\日语单词.xlsx")
sn = wb.sheetnames[0]
ws = wb[sn]
rows = list(ws.iter_rows(values_only=True))
header = rows[0]
df = pd.DataFrame(rows[1:], columns=header)
df.rename(columns={
    header[0]: 'kana',
    header[1]: 'reading',
    header[2]: 'kanji',
    header[3]: 'chinese',
    header[4]: 'note',
}, inplace=True)

# ─── D. reading 字段分析 ────────────────────────────
print("=" * 70)
print("【D】reading 字段取值分析")
print("=" * 70)
vals = df['reading'].dropna().apply(lambda x: str(x)).unique()
print(f"  唯一值 (共{len(vals)}种): {sorted(vals)}")

# reading 为非 0/1 的行
invalid_r = df[df['reading'].apply(
    lambda x: x not in (0.0, 1.0, None) and pd.notna(x)
)]
print(f"\n  reading 不是 0 或 1 的行数: {len(invalid_r)}")
if len(invalid_r):
    print(invalid_r[['kana','reading','kanji','chinese']].to_string())

# ─── E. chinese 字段是数字（应为文字释义）─────────────────
print("\n" + "=" * 70)
print("【E】chinese 字段内容为数字的行（明显错误）")
print("=" * 70)
# 检测 chinese 为纯数字或浮点数
def is_numeric(val):
    if pd.isna(val):
        return False
    try:
        float(str(val).strip())
        return True
    except:
        return False

numeric_cn = df[df['chinese'].apply(is_numeric)]
print(f"  chinese 为纯数字的行数: {len(numeric_cn)}")
if len(numeric_cn):
    print(numeric_cn[['kana','reading','kanji','chinese']].to_string())

# ─── F. chinese 为空值 ──────────────────────────────
print("\n" + "=" * 70)
print("【F】chinese 为空（None/NaN）的行")
print("=" * 70)
cn_null = df[df['chinese'].isna()]
print(f"  chinese 为空行数: {len(cn_null)}")
if len(cn_null):
    print(cn_null[['kana','reading','kanji','chinese']].to_string())

# ─── G. reading 为数字（非 0/1）─────────────────────
print("\n" + "=" * 70)
print("【G】reading 字段所有非 0/1/None 的值（详细）")
print("=" * 70)
all_reading_vals = df['reading'].dropna().tolist()
non_01 = [v for v in all_reading_vals if v not in (0.0, 1.0)]
print(f"  非 0/1 的 reading 值（前50个）: {sorted(set(str(v) for v in non_01))[:50]}")

# ─── H. #NAME? 错误 ─────────────────────────────────
print("\n" + "=" * 70)
print("【H】包含 #NAME? 等 Excel 公式错误的行")
print("=" * 70)
error_mask = (
    df['kana'].astype(str).str.contains(r'#NAME\?|#REF!|#VALUE!|#DIV/0!', na=False) |
    df['kanji'].astype(str).str.contains(r'#NAME\?|#REF!|#VALUE!|#DIV/0!', na=False) |
    df['chinese'].astype(str).str.contains(r'#NAME\?|#REF!|#VALUE!|#DIV/0!', na=False)
)
print(f"  含公式错误的行数: {int(error_mask.sum())}")
if int(error_mask.sum()):
    print(df[error_mask][['kana','reading','kanji','chinese']].to_string())

# ─── I. 用 JSON 比对：chinese 字段与 JSON 不一致 ─────────────
print("\n" + "=" * 70)
print("【I】用 word_data.json 比对：chinese 字段差异")
print("=" * 70)
with open(f"{BASE}\\word_data.json", "r", encoding="utf-8") as f:
    jdata = json.load(f)
jdf = pd.DataFrame(jdata)

# 以 (japanese, kanji) 为 key，比较 chinese
excel_map = {}
for _, row in df.iterrows():
    k = (str(row['kana']) if pd.notna(row['kana']) else '',
         str(row['kanji']) if pd.notna(row['kanji']) else '')
    excel_map[k] = str(row['chinese']) if pd.notna(row['chinese']) else ''

json_map = {}
for _, row in jdf.iterrows():
    k = (str(row['japanese']) if pd.notna(row['japanese']) else '',
         str(row['kanji']) if pd.notna(row['kanji']) else '')
    json_map[k] = str(row['chinese']) if pd.notna(row['chinese']) else ''

diff_cn = []
for k in set(excel_map.keys()) & set(json_map.keys()):
    ev = excel_map[k]
    jv = json_map[k]
    if ev != jv:
        diff_cn.append((k[0], k[1], ev, jv))

print(f"  Excel 与 JSON chinese 字段不一致的条目数: {len(diff_cn)}")
for kana, kanji, ev, jv in diff_cn[:30]:
    print(f"    kana={kana}, kanji={kanji}")
    print(f"      Excel: {ev}")
    print(f"      JSON:  {jv}")

# ─── J. 统计各 reading 值对应的词性含义 ─────────────────
print("\n" + "=" * 70)
print("【J】reading 字段含义解析（按值统计）")
print("=" * 70)
from collections import Counter
rc = Counter(df['reading'].tolist())
for val, cnt in sorted(rc.items(), key=lambda x: (x[0] is None, x[0])):
    sample = df[df['reading'] == val]['kana'].head(5).tolist() if val is not None else \
             df[df['reading'].isna()]['kana'].head(5).tolist()
    print(f"  reading={val}: {cnt} 条 | 示例: {sample}")

print("\n" + "=" * 70)
print("检查完成")
print("=" * 70)
